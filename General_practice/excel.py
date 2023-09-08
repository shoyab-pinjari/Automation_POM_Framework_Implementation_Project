from selenium import webdriver
import openpyxl

file="D:\\Users\Admin\Desktop\IMPORTANT\prepared\lat\TestData.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

rows = sheet.max_row
columns = sheet.max_column

for r in range(2,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(r,c).value, end="      ")
    print()
